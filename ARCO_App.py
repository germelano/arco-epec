"""
ARCO — Análisis y Revisión de Curvas de Obra  v2.0
==================================================
Aplicación web · Streamlit
"""

import io, calendar, math
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════
MONTH_IDX = 10
ORANGE    = "FFFFA500"

MESES_ES  = {"enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
             "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12}
MESES_REV = {v:k for k,v in MESES_ES.items()}

MOTIVOS = [
    "Lluvia / Anegamiento",
    "Permisos",
    "Interferencias",
    "Demora respuesta técnica EPEC",
    "Demora corte de energía",
    "Modificación de proyecto EPEC",
    "Conflicto gremial",
    "Falta de acceso al predio",
    "Condiciones climáticas extremas",
]
MOTIVOS_USD = {"Demora respuesta técnica EPEC", "Modificación de proyecto EPEC"}

# ══════════════════════════════════════════════════════════════════════════════
#  ESTILOS EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def fill(h): return PatternFill("solid", fgColor=h)
_thin  = Side(style="thin",   color="BFBFBF")
_thick = Side(style="medium", color="595959")
B      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
B_SEP  = Border(left=_thin, right=_thin, top=_thin, bottom=_thick)
AC     = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FILLS = {
    "hdr":      fill("1F3864"),
    "actual":   fill("D9E1F2"),
    "contrat":  fill("E2EFDA"),
    "epec":     fill("EDE7F6"),
    "cambiado": fill("FFD966"),
    "subio":    fill("C6EFCE"),
    "bajo":     fill("FFC7CE"),
    "dpos":     fill("A9D18E"),
    "dneg":     fill("FF7B7B"),
    "dbg":      fill("F2F2F2"),
    "lact":     fill("2E75B6"),
    "lcont":    fill("70AD47"),
    "lepec":    fill("7030A0"),
    "ldelta":   fill("595959"),
    "ok_b":     fill("70AD47"),
    "warn_b":   fill("C00000"),
    "ok_row":   fill("C6EFCE"),
    "viol_row": fill("FFC7CE"),
}

def F(bold=False, sz=8, color="000000"):
    return Font(name="Calibri", bold=bold, size=sz, color=color)

# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES DE FECHA
# ══════════════════════════════════════════════════════════════════════════════
def parse_hdr(name):
    try:
        p = str(name).lower().split()
        return (MESES_ES[p[2]], int(p[4]))
    except Exception:
        return None

def days_in(name):
    p = parse_hdr(name)
    return calendar.monthrange(p[1], p[0])[1] if p else 30

def next_month_hdr(name):
    p = parse_hdr(name)
    if not p: return None
    m, y = p
    m += 1
    if m > 12: m, y = 1, y + 1
    d = calendar.monthrange(y, m)[1]
    return f"{d} de {MESES_REV[m]} de {y}"

def _short_month(full_name):
    p = parse_hdr(full_name)
    if not p: return full_name
    m, y = p
    return f"{MESES_REV[m].capitalize()} {y}  ({calendar.monthrange(y,m)[1]}d)"

# ══════════════════════════════════════════════════════════════════════════════
#  CLASIFICACIÓN
# ══════════════════════════════════════════════════════════════════════════════
def classify(nro, desc, moneda=""):
    d = str(desc or "")
    m = str(moneda or "").strip().upper()
    if "U$" in m or "USD" in m or "U$D" in d or "U$d" in d:
        return "USD"
    if str(nro) in ("1,01", "2,01") and "CONFORME" not in d.upper():
        return "DOC"
    return "CE"

def is_affected(nro, desc, motive, moneda=""):
    k = classify(nro, desc, moneda)
    if k == "DOC":  return False
    if k == "USD":  return motive in MOTIVOS_USD
    return True

# ══════════════════════════════════════════════════════════════════════════════
#  LECTURA (acepta ruta string o BytesIO)
# ══════════════════════════════════════════════════════════════════════════════
def _detect_header_row(ws):
    """Devuelve la fila del encabezado (1 o 2) detectando si hay fila de descripción."""
    row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    val = row1[MONTH_IDX] if len(row1) > MONTH_IDX else None
    return 1 if parse_hdr(val) else 2

def read_file(src):
    if isinstance(src, bytes):
        src = io.BytesIO(src)
    wb = load_workbook(src, data_only=True)
    ws = wb["Hoja1"]
    hdr_row = _detect_header_row(ws)
    header = [c.value for c in next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row))]
    items  = [list(r) for r in ws.iter_rows(min_row=hdr_row+1, values_only=True) if r[0]]
    return header, items

def read_orange(src):
    if isinstance(src, bytes):
        src = io.BytesIO(src)
    wb = load_workbook(src, data_only=True)
    ws = wb["Hoja1"]
    data_start = _detect_header_row(ws) + 1
    result = {}
    for row in ws.iter_rows(min_row=data_start):
        if not row[0].value: continue
        rid = row[0].value
        for cell in row:
            if cell.column < MONTH_IDX + 1 or not cell.value: continue
            fc = cell.font.color if cell.font else None
            if fc and fc.type == "rgb" and fc.rgb == ORANGE:
                result.setdefault(rid, {})[cell.column] = cell.value
    return result

def read_computo(src):
    """Lee el cómputo métrico → {nro: pct_acumulado (0-1)}
    Soporta el formato SIGO que tiene dos hojas: 'Resumen' y 'DetalleComputo'.
    Lee 'DetalleComputo' si existe; si no, cae al sheet activo.
    Columnas en DetalleComputo (fila 2 = encabezado, fila 3+ = datos):
      col 1 (idx 0): Nro ítem
      col 10 (idx 9): Cantidad medición acumulada  ← anterior + actual, fracción 0-1
    """
    if isinstance(src, bytes):
        src = io.BytesIO(src)
    wb = load_workbook(src, data_only=True)
    ws = wb["DetalleComputo"] if "DetalleComputo" in wb.sheetnames else wb.active
    result = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        if len(row) <= 9: continue   # fila con menos columnas de las esperadas
        nro = str(row[0]).strip()
        pct = float(row[9] or 0)   # Cantidad medición acumulada = anterior + actual (fracción 0-1)
        result[nro] = pct
    return result

def build_computo_floors(items, computo):
    floors = {}
    for item in items:
        iid = item[0]
        nro = str(item[2] or '').strip()
        if not computo or nro not in computo:
            floors[iid] = {}
            continue
        pct_cert = computo[nro]
        actuals = [float(v or 0) for v in item[MONTH_IDX:]]
        cumsum = 0.0
        item_floors = {}
        for mi, val in enumerate(actuals):
            if val <= 0:
                continue
            remaining = pct_cert - cumsum
            if remaining <= 1e-9:
                break  # ya certificamos todo
            certified = min(val, remaining)
            item_floors[mi] = certified
            cumsum += val
        floors[iid] = item_floors
    return floors

# ══════════════════════════════════════════════════════════════════════════════
#  ALGORITMO DE CORRIMIENTO
# ══════════════════════════════════════════════════════════════════════════════
def shift_one(pcts, idx, fraction, item_max, certified_val=0.0):
    pcts = list(pcts)
    shiftable = max(0.0, pcts[idx] - certified_val)
    amount = shiftable * fraction
    if amount < 1e-10:
        return pcts, 0.0
    pcts[idx] -= amount
    remaining  = amount
    t = idx + 1
    while remaining > 1e-10:
        if t >= len(pcts):
            return pcts, remaining
        avail = item_max - pcts[t]
        if avail > 1e-10:
            absorb    = min(remaining, avail)
            pcts[t]  += absorb
            remaining -= absorb
        t += 1
    return pcts, 0.0

def apply_delays(items, header, delays, naranja, extra_months=0, computo_floors=None):
    month_names = list(header[MONTH_IDX:])
    for _ in range(extra_months):
        month_names.append(next_month_hdr(month_names[-1]))
    new_header  = list(header[:MONTH_IDX]) + month_names
    m_to_idx    = {n: i for i, n in enumerate(month_names)}
    result        = []
    overflow_nros = []
    for item in items:
        iid    = item[0]; nro = item[2]; desc = item[3]
        moneda = item[5] if len(item) > 5 else ""
        base = [float(v or 0) for v in item[MONTH_IDX:]]
        orig = list(base)
        pcts = base + [0.0] * extra_months
        item_max        = max((v for v in orig if v and v > 0), default=1.0)
        item_orange     = naranja.get(iid, {})
        item_comp_floor = (computo_floors or {}).get(iid, {})
        total_ovf       = 0.0
        for d in delays:
            if not is_affected(nro, desc, d["motive"], moneda): continue
            mn = d["month_name"]
            if mn not in m_to_idx: continue
            mi   = m_to_idx[mn]
            days = min(int(d["days"]), days_in(mn))
            frac = days / days_in(mn)
            col1 = MONTH_IDX + 1 + mi
            cert_naranja = float(item_orange.get(col1, 0))
            cert_computo = item_comp_floor.get(mi, 0)
            cert = max(cert_naranja, cert_computo)
            pcts, ovf = shift_one(pcts, mi, frac, item_max, cert)
            total_ovf += ovf
        if total_ovf > 1e-10:
            overflow_nros.append(nro)
        new_row = list(item[:MONTH_IDX]) + [
            round(p, 6) if p > 1e-10 else None for p in pcts
        ]
        result.append(new_row)
    return result, new_header, overflow_nros

def compress_items(items, overflow_nros):
    ovf_set = set(overflow_nros)
    result  = []
    for item in items:
        if item[2] not in ovf_set:
            result.append(item); continue
        pcts    = [float(v or 0) for v in item[MONTH_IDX:]]
        missing = 1.0 - sum(pcts)
        if missing > 1e-10:
            pos_sum = sum(p for p in pcts if p > 0)
            if pos_sum > 0:
                pcts = [p + missing*(p/pos_sum) if p > 0 else p for p in pcts]
        new_row = list(item[:MONTH_IDX]) + [
            round(p, 6) if p > 1e-10 else None for p in pcts
        ]
        result.append(new_row)
    return result

# ══════════════════════════════════════════════════════════════════════════════
#  GENERACIÓN DE EXCEL  (devuelve bytes)
# ══════════════════════════════════════════════════════════════════════════════
DESC_COL_WIDTH = 48

def row_height(desc, font_size=8):
    if not desc: return 15
    chars = len(str(desc))
    lines = max(1, math.ceil(chars / (DESC_COL_WIDTH * 1.1)))
    return max(15, int(lines * font_size * 1.7) + 4)

DIFF_THRESHOLD = 0.005

def _diff_map(a_vals, b_vals):
    out = {}
    for i, (a, b) in enumerate(zip(a_vals, b_vals)):
        fa = round(float(a or 0), 6); fb = round(float(b or 0), 6)
        if abs(fb - fa) < DIFF_THRESHOLD: continue
        if   fa == 0: out[i] = "cambiado"
        elif fb == 0: out[i] = "cambiado"
        elif fb > fa: out[i] = "subio"
        else:         out[i] = "bajo"
    return out

def _write_item_row(ws, row, nro, desc, mon, prec, lbl, lbl_fill,
                    month_vals, base_fill, diff=None, sep=False):
    bdr = B_SEP if sep else B
    for ci, val in enumerate([nro, desc, mon, prec, lbl] + list(month_vals), 1):
        c = ws.cell(row, ci, val); c.border = bdr
        if ci == 5:
            c.fill = lbl_fill; c.font = F(True,8,"FFFFFF"); c.alignment = AC
        elif ci <= 5:
            c.fill = base_fill
            c.font = F(True,8) if ci in (1,2) else F()
            c.alignment = AL if ci == 2 else AC
            if ci == 4 and val: c.number_format = "#,##0.00"
        else:
            midx = ci - 6; c.alignment = AC; c.font = F()
            if val is not None: c.number_format = "0.00%"
            tag = (diff or {}).get(midx)
            c.fill = FILLS[tag] if tag else base_fill
            if tag: c.font = F(True, 8)
    ws.row_dimensions[row].height = row_height(desc)

def _write_delta_row(ws, row, nro, vals_a, vals_b, n_meses):
    for ci in range(1, 6 + n_meses + 1):
        c = ws.cell(row, ci); c.border = B_SEP
        if ci == 5:
            c.value = "DIFERENCIA"; c.fill = FILLS["ldelta"]
            c.font = F(True,8,"FFFFFF"); c.alignment = AC
        elif ci < 5:
            c.fill = FILLS["dbg"]
            if ci == 1: c.value = nro; c.font = F(); c.alignment = AC
        else:
            idx = ci - 6
            va  = float(vals_a[idx] or 0) if idx < len(vals_a) else 0
            vb  = float(vals_b[idx] or 0) if idx < len(vals_b) else 0
            d   = round(vb - va, 6)
            c.alignment = AC; c.font = F(True, 8)
            if abs(d) >= DIFF_THRESHOLD:
                c.value = d; c.number_format = '+0.00%;-0.00%;""'
                c.fill  = FILLS["dpos"] if d > 0 else FILLS["dneg"]
            else:
                c.fill = FILLS["dbg"]
    ws.row_dimensions[row].height = row_height(nro)


def generate_excel(src_actual, src_contrat, epec_items, epec_header,
                   naranja, obra_name="", computo=None):
    """Genera el Excel y devuelve bytes."""
    hdr_a, items_a = read_file(src_actual)
    items_c = None
    if src_contrat is not None:
        _, items_c = read_file(src_contrat)

    prop_c = {r[0]: r for r in items_c}    if items_c   else {}
    prop_e = {r[0]: r for r in epec_items} if epec_items else {}

    use_hdr   = epec_header if epec_header else hdr_a
    meses     = [h for h in use_hdr[MONTH_IDX:] if h]
    N         = len(meses)
    TOTAL_COL = 5 + N
    prefix    = f"{obra_name}  —  " if obra_name else ""

    wb = Workbook()

    # ── HOJA 1: COMPARATIVO ───────────────────────────────────────────────────
    ws = wb.active; ws.title = "Comparativo"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COL)
    t = ws.cell(1, 1, f"{prefix}ARCO  |  CURVA DE AVANCE  |  ACTUAL vs PROPUESTAS")
    t.fill = FILLS["hdr"]; t.alignment = AC
    t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws.row_dimensions[1].height = 22

    leyenda = [("ACTUAL", FILLS["lact"], "FFFFFF"),
               ("PROPUESTA CONTRATISTA", FILLS["lcont"], "FFFFFF"),
               ("PROPUESTA EPEC", FILLS["lepec"], "FFFFFF"),
               ("↑ SUBIÓ", FILLS["subio"], "000000"),
               ("↓ BAJÓ", FILLS["bajo"], "000000"),
               ("NUEVO / QUITADO", FILLS["cambiado"], "000000")]
    col = 6
    for txt, flt, fc in leyenda:
        c = ws.cell(2, col, txt); c.fill = flt; c.border = B; c.alignment = AC
        c.font = Font(name="Calibri", bold=True, size=8, color=fc); col += 3
    ws.row_dimensions[2].height = 22

    for ci, h in enumerate(["Nro","Descripción","Mon.","Precio Oficial","Tipo"] + meses, 1):
        c = ws.cell(3, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
    ws.row_dimensions[3].height = 52

    ws.column_dimensions["A"].width = 7;  ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 5;  ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 13
    for i in range(N):
        ws.column_dimensions[get_column_letter(6+i)].width = 7

    cur = 4; cambios_c = []; cambios_e = []
    for item_a in items_a:
        iid  = item_a[0]; nro = item_a[2]; desc = item_a[3]
        mon  = item_a[5]; prec = item_a[9]
        ma   = [item_a[MONTH_IDX+i] if MONTH_IDX+i < len(item_a) else None for i in range(N)]
        rc   = prop_c.get(iid)
        mc   = [rc[MONTH_IDX+i] if rc and MONTH_IDX+i < len(rc) else None for i in range(N)]
        re   = prop_e.get(iid)
        me   = [re[MONTH_IDX+i] if re and MONTH_IDX+i < len(re) else None for i in range(N)]

        if rc and any(round(float(a or 0),6) != round(float(c or 0),6) for a,c in zip(ma,mc)):
            cambios_c.append((nro, desc))
        if re and any(round(float(a or 0),6) != round(float(e or 0),6) for a,e in zip(ma,me)):
            cambios_e.append((nro, desc))

        _write_item_row(ws, cur, nro, desc, mon, prec,
                        "ACTUAL", FILLS["lact"], ma, FILLS["actual"]); cur += 1
        if rc:
            _write_item_row(ws, cur, nro, desc, mon, prec,
                            "CONTRATISTA", FILLS["lcont"], mc, FILLS["contrat"], _diff_map(ma,mc))
            cur += 1
            _write_delta_row(ws, cur, nro, ma, mc, N); cur += 1
        if re:
            _write_item_row(ws, cur, nro, desc, mon, prec,
                            "PROPUESTA EPEC", FILLS["lepec"], me, FILLS["epec"], _diff_map(ma,me),
                            sep=not rc)
            cur += 1
            _write_delta_row(ws, cur, nro, ma, me, N); cur += 1
        if not rc and not re:
            for ci2 in range(1, TOTAL_COL+1):
                ws.cell(cur-1, ci2).border = B_SEP
    ws.freeze_panes = ws.cell(4, 6)

    # ── HOJA: EPEC vs CONTRATISTA ─────────────────────────────────────────────
    if prop_c and prop_e:
        ws_ce = wb.create_sheet("EPEC vs Contratista")
        ws_ce.column_dimensions["A"].width = 7;  ws_ce.column_dimensions["B"].width = 48
        ws_ce.column_dimensions["C"].width = 5;  ws_ce.column_dimensions["D"].width = 14
        ws_ce.column_dimensions["E"].width = 13
        for i in range(N):
            ws_ce.column_dimensions[get_column_letter(6+i)].width = 7

        ws_ce.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COL)
        t_ce = ws_ce.cell(1, 1, f"{prefix}ARCO  |  PROPUESTA EPEC vs PROPUESTA CONTRATISTA")
        t_ce.fill = FILLS["hdr"]; t_ce.alignment = AC
        t_ce.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        ws_ce.row_dimensions[1].height = 22

        for txt, flt, fc in [("PROPUESTA EPEC",FILLS["lepec"],"FFFFFF"),
                              ("PROPUESTA CONTRATISTA",FILLS["lcont"],"FFFFFF"),
                              ("CONTRAT. PIDE MÁS",FILLS["bajo"],"000000"),
                              ("CONTRAT. PIDE MENOS",FILLS["subio"],"000000"),
                              ("NUEVO / QUITADO",FILLS["cambiado"],"000000")]:
            c = ws_ce.cell(2, col, txt); c.fill = flt; c.border = B; c.alignment = AC
            c.font = Font(name="Calibri", bold=True, size=8, color=fc); col += 3
        ws_ce.row_dimensions[2].height = 22

        for ci, h in enumerate(["Nro","Descripción","Mon.","Precio Oficial","Tipo"] + meses, 1):
            c = ws_ce.cell(3, ci, h); c.fill = FILLS["hdr"]
            c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
        ws_ce.row_dimensions[3].height = 52

        cur_ce = 4
        for item_a in items_a:
            iid = item_a[0]; nro = item_a[2]; desc = item_a[3]
            mon = item_a[5]; prec = item_a[9]
            rc_ce = prop_c.get(iid); re_ce = prop_e.get(iid)
            if not rc_ce and not re_ce: continue
            mc_ce = [rc_ce[MONTH_IDX+i] if rc_ce and MONTH_IDX+i < len(rc_ce) else None for i in range(N)]
            me_ce = [re_ce[MONTH_IDX+i] if re_ce and MONTH_IDX+i < len(re_ce) else None for i in range(N)]
            diff_ce = {}
            for i, (vc2, ve2) in enumerate(zip(mc_ce, me_ce)):
                fc2 = round(float(vc2 or 0), 6); fe2 = round(float(ve2 or 0), 6)
                if abs(fc2 - fe2) < DIFF_THRESHOLD: continue
                if   fc2 == 0: diff_ce[i] = "cambiado"
                elif fe2 == 0: diff_ce[i] = "cambiado"
                elif fc2 > fe2: diff_ce[i] = "bajo"
                else:           diff_ce[i] = "subio"
            _write_item_row(ws_ce, cur_ce, nro, desc, mon, prec,
                            "PROPUESTA EPEC", FILLS["lepec"], me_ce, FILLS["epec"]); cur_ce += 1
            _write_item_row(ws_ce, cur_ce, nro, desc, mon, prec,
                            "CONTRATISTA", FILLS["lcont"], mc_ce, FILLS["contrat"], diff_ce); cur_ce += 1
            _write_delta_row(ws_ce, cur_ce, nro, me_ce, mc_ce, N); cur_ce += 1
        ws_ce.freeze_panes = ws_ce.cell(4, 6)

    # ── HOJA 2: RESUMEN DE CAMBIOS ────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen de Cambios")
    ws2.column_dimensions["A"].width = 8; ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 22
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]; t2.value = f"{prefix}RESUMEN DE CAMBIOS"
    t2.fill = FILLS["hdr"]; t2.alignment = AC
    t2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws2.row_dimensions[1].height = 20
    for ci, h in enumerate(["Nro","Descripción","Propuesta"], 1):
        c = ws2.cell(2, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
    rr = 3
    for nro, desc in cambios_c:
        a=ws2.cell(rr,1,nro); a.border=B; a.alignment=AC; a.font=F()
        b=ws2.cell(rr,2,desc); b.border=B; b.alignment=AL; b.font=F()
        c=ws2.cell(rr,3,"Contratista"); c.border=B; c.alignment=AC
        c.font=F(True,8,"FFFFFF"); c.fill=FILLS["lcont"]
        flt = fill("FFFFFF" if rr%2==0 else "F2F2F2"); a.fill=flt; b.fill=flt
        ws2.row_dimensions[rr].height = row_height(desc); rr+=1
    for nro, desc in cambios_e:
        a=ws2.cell(rr,1,nro); a.border=B; a.alignment=AC; a.font=F()
        b=ws2.cell(rr,2,desc); b.border=B; b.alignment=AL; b.font=F()
        c=ws2.cell(rr,3,"EPEC"); c.border=B; c.alignment=AC
        c.font=F(True,8,"FFFFFF"); c.fill=FILLS["lepec"]
        flt = fill("FFFFFF" if rr%2==0 else "F2F2F2"); a.fill=flt; b.fill=flt
        ws2.row_dimensions[rr].height = row_height(desc); rr+=1

    # ── HOJA 3: VERIFICACIÓN CERTIFICADOS ────────────────────────────────────
    ws3 = wb.create_sheet("Verificación Certificados")
    for col_l, w in zip("ABCDEFGH", [8,52,22,18,18,18,16,14]):
        ws3.column_dimensions[col_l].width = w
    violations = []; ok_list = []

    comp_floors_verify = build_computo_floors(items_a, computo) if computo else {}

    for item_a in items_a:
        iid = item_a[0]; nro = item_a[2]; desc = item_a[3]
        orange_cols = naranja.get(iid, {})
        comp_f      = comp_floors_verify.get(iid, {})

        certified_cols = {}
        for col1, val in orange_cols.items():
            certified_cols[col1] = (val, "SIGO (naranja)")
        for mi, val in comp_f.items():
            col1 = MONTH_IDX + 1 + mi
            if col1 not in certified_cols:
                certified_cols[col1] = (val, "Cómputo")

        if not certified_cols: continue
        for col1, (val_a, fuente) in certified_cols.items():
            mes_name = hdr_a[col1-1] if col1-1 < len(hdr_a) else ""
            rc2 = prop_c.get(iid); re2 = prop_e.get(iid)
            vc  = rc2[col1-1] if rc2 and col1-1 < len(rc2) else None
            ve  = re2[col1-1] if re2 and col1-1 < len(re2) else None
            fa  = round(float(val_a or 0), 8)
            fc2 = round(float(vc or 0), 8) if vc is not None else fa
            fe2 = round(float(ve or 0), 8) if ve is not None else fa
            entry = dict(nro=nro, desc=desc, mes=mes_name, va=val_a, vc=vc, ve=ve, fuente=fuente)
            if fa != fc2 or fa != fe2: violations.append(entry)
            else: ok_list.append(entry)
    ws3.merge_cells("A1:G1"); b1 = ws3["A1"]
    if violations:
        b1.value = f"ALERTA: {len(violations)} celda(s) certificada(s) modificadas"
        b1.fill  = FILLS["warn_b"]
    else:
        b1.value = f"VERIFICACIÓN OK — {len(ok_list)} celdas certificadas sin modificar"
        b1.fill  = FILLS["ok_b"]
    b1.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    b1.alignment = AC; ws3.row_dimensions[1].height = 24
    for ci, h in enumerate(["Nro","Descripción","Mes","Actual","Contratista","EPEC","Estado","Fuente"], 1):
        c = ws3.cell(2, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
    ws3.row_dimensions[2].height = 18
    rv = 3
    for e in violations + ok_list:
        is_v = e in violations
        flt  = FILLS["viol_row"] if is_v else FILLS["ok_row"]
        est  = "MODIFICADO" if is_v else "OK"
        fc_e = Font(name="Calibri", bold=True, size=8, color=("C00000" if is_v else "375623"))
        fuente = e.get("fuente", "SIGO (naranja)")
        for ci, v in enumerate([e["nro"],e["desc"],e["mes"],e["va"],e["vc"],e["ve"],est,fuente], 1):
            c = ws3.cell(rv, ci, v); c.fill = flt; c.border = B
            c.alignment = AL if ci == 2 else AC
            if ci == 7: c.font = fc_e
            elif ci == 8:
                c.font = F(False,8)
                c.fill = FILLS["dbg"] if not is_v else flt
            elif ci in (4,5,6):
                c.font = F(True,8)
                if v is not None: c.number_format = "0.00%"
            else: c.font = F()
        ws3.row_dimensions[rv].height = row_height(e["desc"]); rv += 1
    ws3.freeze_panes = ws3["A3"]

    # ── HOJA: CURVAS ──────────────────────────────────────────────────────────
    def calc_cumul(item_list, n):
        monthly = [0.0] * n
        for item in item_list:
            precio = float(item[9] or 0)
            for i in range(n):
                idx = MONTH_IDX + i
                monthly[i] += precio * (float(item[idx] or 0) if idx < len(item) else 0.0)
        cumul = []; acc = 0.0
        for v in monthly:
            acc += v; cumul.append(round(acc, 2))
        return cumul

    wc = wb.create_sheet("Curvas")
    cumul_a = calc_cumul(items_a, N); total_a = cumul_a[-1] if cumul_a else 1.0
    cumul_c = calc_cumul(list(prop_c.values()), N) if prop_c else None
    cumul_e = calc_cumul(list(prop_e.values()), N) if prop_e else None

    wc.merge_cells("A1:H1"); tc = wc["A1"]
    tc.value = f"{prefix}CURVAS DE AVANCE ACUMULADO"
    tc.fill = FILLS["hdr"]; tc.alignment = AC
    tc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    wc.row_dimensions[1].height = 22

    for ci, h in enumerate(["Mes","Actual ($)","Actual (%)","Propuesta EPEC ($)",
                              "Propuesta EPEC (%)","Contratista ($)","Contratista (%)"], 1):
        c = wc.cell(2, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
    wc.row_dimensions[2].height = 28
    wc.column_dimensions["A"].width = 22
    for col_l in "BCDEFG": wc.column_dimensions[col_l].width = 18

    for i, mes in enumerate(meses):
        r = i + 3; flt = fill("FFFFFF" if i % 2 == 0 else "F2F2F2")
        c = wc.cell(r,1,mes); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC
        va = cumul_a[i]
        c = wc.cell(r,2,va); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC; c.number_format='#,##0'
        c = wc.cell(r,3,va/total_a if total_a else 0); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC; c.number_format='0.00%'
        if cumul_e:
            ve=cumul_e[i]; te=cumul_e[-1] if cumul_e[-1] else 1.0
            c=wc.cell(r,4,ve); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC; c.number_format='#,##0'
            c=wc.cell(r,5,ve/te); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC; c.number_format='0.00%'
        if cumul_c:
            vc=cumul_c[i]; tc2=cumul_c[-1] if cumul_c[-1] else 1.0
            c=wc.cell(r,6,vc); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC; c.number_format='#,##0'
            c=wc.cell(r,7,vc/tc2); c.fill=flt; c.border=B; c.font=F(); c.alignment=AC; c.number_format='0.00%'
        wc.row_dimensions[r].height = 15

    chart = LineChart()
    chart.title = "Curvas de Avance Acumulado"; chart.style = 10
    chart.height = 18; chart.width = 32
    max_val = max(cumul_a) if cumul_a else 1
    if max_val >= 1_000_000_000:
        chart.y_axis.numFmt = '#,##0,,"M"'; chart.y_axis.title = "Monto acumulado (en millones $)"
    elif max_val >= 1_000_000:
        chart.y_axis.numFmt = '#,##0,"k"'; chart.y_axis.title = "Monto acumulado (en miles $)"
    else:
        chart.y_axis.numFmt = '#,##0'; chart.y_axis.title = "Monto acumulado ($)"
    chart.y_axis.delete = False; chart.y_axis.crosses = "autoZero"; chart.y_axis.crossAx = 100
    chart.x_axis.title = "Mes"; chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1; chart.x_axis.noMultiLvlLbl = True

    cats = Reference(wc, min_col=1, min_row=3, max_row=2+N); chart.set_categories(cats)
    ser_a = Reference(wc, min_col=2, min_row=2, max_row=2+N); chart.add_data(ser_a, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chart.series[0].graphicalProperties.line.width     = 20000
    chart.series[0].smooth = True
    if cumul_e:
        ser_e = Reference(wc, min_col=4, min_row=2, max_row=2+N); chart.add_data(ser_e, titles_from_data=True)
        chart.series[-1].graphicalProperties.line.solidFill = "7030A0"
        chart.series[-1].graphicalProperties.line.width     = 20000
        chart.series[-1].smooth = True
    if cumul_c:
        ser_c = Reference(wc, min_col=6, min_row=2, max_row=2+N); chart.add_data(ser_c, titles_from_data=True)
        chart.series[-1].graphicalProperties.line.solidFill = "70AD47"
        chart.series[-1].graphicalProperties.line.width     = 20000
        chart.series[-1].smooth = True
    wc.add_chart(chart, f"A{N+5}")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  INTERFAZ STREAMLIT
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="ARCO — EPEC", page_icon="📊", layout="centered")

st.markdown("""
<div style='background:#1F3864;padding:20px 24px;border-radius:8px;margin-bottom:24px'>
  <h1 style='color:white;margin:0;font-size:28px;font-family:Calibri,sans-serif'>ARCO</h1>
  <p style='color:#AAC4E8;margin:6px 0 0 0;font-size:14px;font-family:Calibri,sans-serif'>
    Análisis y Revisión de Curvas de Obra &nbsp;·&nbsp; EPEC
  </p>
</div>
""", unsafe_allow_html=True)

# ── IDENTIFICACIÓN ────────────────────────────────────────────────────────────
st.subheader("Identificación de obra")
obra_name = st.text_input("Nombre de la obra", placeholder="Ej: ET Los Espinillos")
st.caption("Aparece en el título del Excel y en el nombre del archivo generado.")

st.divider()

# ── ARCHIVOS ─────────────────────────────────────────────────────────────────
st.subheader("Archivos")
col1, col2 = st.columns(2)
with col1:
    uploaded_actual = st.file_uploader(
        "📘 Curva Actual (requerida)", type=["xlsx","xlsm"],
        help="Descargada desde SIGO, sin modificar.")
with col2:
    uploaded_contrat = st.file_uploader(
        "📗 Propuesta Contratista (opcional)", type=["xlsx","xlsm"],
        help="Mismo formato que la Curva Actual.")

uploaded_computo = st.file_uploader(
    "📊 Cómputo Métrico — protege lo ya certificado en todos los modos",
    type=["xlsx"],
    help="Exportado desde SIGO (Ver cómputo). Usar el del último mes disponible.")

months_full = []; months_display = []
if uploaded_actual:
    try:
        hdr_preview, _ = read_file(io.BytesIO(uploaded_actual.getvalue()))
        months_full    = [h for h in hdr_preview[MONTH_IDX:] if h]
        months_display = [_short_month(m) for m in months_full]
        st.success(f"✓ Curva Actual cargada — {len(months_full)} meses detectados.")
    except Exception as ex:
        st.error(f"Error al leer la Curva Actual: {ex}")

st.divider()

# ── PROPUESTA EPEC ────────────────────────────────────────────────────────────
st.subheader("Propuesta EPEC")
use_epec = st.checkbox("Generar Propuesta EPEC (prórroga por motivos)")

delays = []
overflow_mode = "Ampliación de plazo"

if use_epec:
    if not uploaded_actual:
        st.warning("Cargá primero el archivo de Curva Actual.")
    else:
        if "n_delays" not in st.session_state:
            st.session_state.n_delays = 1

        st.caption("Los ítems en USD solo se mueven por _Demora respuesta técnica EPEC_ y _Modificación de proyecto EPEC_.")

        h1, h2, h3, _ = st.columns([3, 2.5, 1, 0.5])
        h1.markdown("**Motivo**"); h2.markdown("**Mes**"); h3.markdown("**Días**")

        to_delete = None
        for i in range(st.session_state.n_delays):
            c1, c2, c3, c4 = st.columns([3, 2.5, 1, 0.5])
            with c1:
                motive = st.selectbox("", MOTIVOS, key=f"mot_{i}", label_visibility="collapsed")
            with c2:
                opts = months_display if months_display else ["(sin meses)"]
                disp = st.selectbox("", opts, key=f"mes_{i}", label_visibility="collapsed")
                month_full = months_full[months_display.index(disp)] if months_display else ""
            with c3:
                dias = st.number_input("", min_value=1, max_value=31, value=15,
                                       key=f"dias_{i}", label_visibility="collapsed")
            with c4:
                if st.session_state.n_delays > 1:
                    if st.button("✕", key=f"del_{i}"):
                        to_delete = i
            delays.append({"motive": motive, "month_name": month_full, "days": int(dias)})

        if to_delete is not None:
            st.session_state.n_delays -= 1
            st.rerun()

        if st.button("＋ Agregar motivo"):
            st.session_state.n_delays += 1
            st.rerun()

        st.divider()
        overflow_mode = st.radio(
            "Seleccioná el tipo de modificación a aplicar:",
            ["Ampliación de plazo", "Modificación de curva (sin extender plazo)"],
            horizontal=True,
            help="Ampliación de plazo agrega meses al final del proyecto. Modificación de curva redistribuye las actividades restantes dentro del plazo actual.")

st.divider()

# ── BOTÓN GENERAR ─────────────────────────────────────────────────────────────
can_generate = (
    uploaded_actual is not None and
    uploaded_computo is not None and
    (uploaded_contrat is not None or use_epec) and
    (not use_epec or len(delays) > 0)
)

if uploaded_actual and not uploaded_computo:
    st.info("Cargá el Cómputo Métrico para continuar.")
elif uploaded_actual and not uploaded_contrat and not use_epec:
    st.info("Cargá la Propuesta del Contratista o activá la Propuesta EPEC para continuar.")

if st.button("▶  GENERAR", disabled=not can_generate, type="primary", use_container_width=True):
    with st.spinner("Procesando..."):
        try:
            actual_bytes   = uploaded_actual.getvalue()
            contrat_bytes  = uploaded_contrat.getvalue() if uploaded_contrat else None
            computo_bytes  = uploaded_computo.getvalue() if uploaded_computo else None

            hdr_a, items_a = read_file(io.BytesIO(actual_bytes))
            naranja        = read_orange(io.BytesIO(actual_bytes))
            computo        = read_computo(io.BytesIO(computo_bytes)) if computo_bytes else None
            comp_floors    = build_computo_floors(items_a, computo) if computo else None

            ep_items = None; ep_header = None; overflow_warning = False

            if use_epec and delays:
                ep_items, ep_hdr, ovf = apply_delays(
                    items_a, hdr_a, delays, naranja, extra_months=0,
                    computo_floors=comp_floors)
                if ovf:
                    overflow_warning = True
                    if overflow_mode == "Ampliación de plazo":
                        for extra in [2, 4, 6]:
                            ep_items, ep_hdr, ovf = apply_delays(
                                items_a, hdr_a, delays, naranja, extra_months=extra,
                                computo_floors=comp_floors)
                            if not ovf: break
                    else:
                        ep_items = compress_items(ep_items, ovf)
                ep_header = ep_hdr

            excel_bytes = generate_excel(
                src_actual   = io.BytesIO(actual_bytes),
                src_contrat  = io.BytesIO(contrat_bytes) if contrat_bytes else None,
                epec_items   = ep_items,
                epec_header  = ep_header,
                naranja      = naranja,
                obra_name    = obra_name.strip(),
                computo      = computo,
            )

            safe  = "".join(c if c.isalnum() or c in "_-" else "_"
                            for c in obra_name.strip()).strip("_")
            fname = f"ARCO_{safe}.xlsx" if safe else "ARCO_Curva_Avance.xlsx"

            if overflow_warning:
                msg = "se extendió el proyecto" if overflow_mode == "Ampliación de plazo" \
                      else "se comprimieron las actividades finales"
                st.warning(f"⚠️ Algunos ítems superaban el plazo original — {msg}.")

            st.success("✅ ¡Excel generado correctamente!")
            st.download_button(
                label=f"📥 Descargar {fname}",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as ex:
            import traceback
            st.error(f"Error al generar: {ex}")
            with st.expander("Detalle técnico"):
                st.code(traceback.format_exc())

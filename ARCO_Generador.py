"""
ARCO — Análisis y Revisión de Curvas de Obra  v2.0
==================================================
Modo 1: Compara Curva Actual vs Propuesta del Contratista
Modo 2: Genera Propuesta EPEC por motivos de prórroga
Ambos modos pueden usarse juntos en el mismo Excel de salida.
"""

import sys, os, calendar, threading
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl.chart import LineChart, Reference

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════
MONTH_IDX = 10   # índice 0-based de la primera columna de mes en cada fila
ORANGE    = "FFFFA500"

MESES_ES  = {"enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
             "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12}
MESES_REV = {v:k for k,v in MESES_ES.items()}

MOTIVOS = [
    "Lluvia / Anegamiento",
    "Permisos",
    "Interferencias",
    "Demora respuesta técnica EPEC",
    "Demora corte de energía",
    "Modificación de proyecto EPEC",
    "Conflicto gremial",
    "Falta de acceso al predio",
    "Condiciones climáticas extremas",
]
# Motivos que también afectan ítems en USD
MOTIVOS_USD = {"Demora respuesta técnica EPEC", "Modificación de proyecto EPEC"}

# ══════════════════════════════════════════════════════════════════════════════
#  ESTILOS EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def fill(h): return PatternFill("solid", fgColor=h)
_thin  = Side(style="thin",   color="BFBFBF")
_thick = Side(style="medium", color="595959")
B      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
B_SEP  = Border(left=_thin, right=_thin, top=_thin, bottom=_thick)
AC     = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FILLS = {
    "hdr":      fill("1F3864"),
    "actual":   fill("D9E1F2"),
    "contrat":  fill("E2EFDA"),
    "epec":     fill("EDE7F6"),
    "cambiado": fill("FFD966"),
    "subio":    fill("C6EFCE"),
    "bajo":     fill("FFC7CE"),
    "dpos":     fill("A9D18E"),
    "dneg":     fill("FF7B7B"),
    "dbg":      fill("F2F2F2"),
    "lact":     fill("2E75B6"),
    "lcont":    fill("70AD47"),
    "lepec":    fill("7030A0"),
    "ldelta":   fill("595959"),
    "ok_b":     fill("70AD47"),
    "warn_b":   fill("C00000"),
    "ok_row":   fill("C6EFCE"),
    "viol_row": fill("FFC7CE"),
}

def F(bold=False, sz=8, color="000000"):
    return Font(name="Calibri", bold=bold, size=sz, color=color)

# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES DE FECHA
# ══════════════════════════════════════════════════════════════════════════════
def parse_hdr(name):
    """'31 de marzo de 2025' → (3, 2025)  |  None si no parsea"""
    try:
        p = str(name).lower().split()
        return (MESES_ES[p[2]], int(p[4]))
    except Exception:
        return None

def days_in(name):
    p = parse_hdr(name)
    return calendar.monthrange(p[1], p[0])[1] if p else 30

def next_month_hdr(name):
    p = parse_hdr(name)
    if not p: return None
    m, y = p
    m += 1
    if m > 12: m, y = 1, y + 1
    d = calendar.monthrange(y, m)[1]
    return f"{d} de {MESES_REV[m]} de {y}"

# ══════════════════════════════════════════════════════════════════════════════
#  CLASIFICACIÓN DE ÍTEMS
# ══════════════════════════════════════════════════════════════════════════════
def classify(nro, desc):
    """Devuelve 'USD' | 'DOC' | 'CE' (civil/electromecánico)"""
    d = desc or ""
    if "U$D" in d or "U$d" in d:
        return "USD"
    # Doc técnica que NO es conforme a obra → no se mueve
    if str(nro) in ("1,01", "2,01") and "CONFORME" not in d.upper():
        return "DOC"
    return "CE"

def is_affected(nro, desc, motive):
    k = classify(nro, desc)
    if k == "DOC":  return False
    if k == "USD":  return motive in MOTIVOS_USD
    return True  # CE siempre afectado

# ══════════════════════════════════════════════════════════════════════════════
#  LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════════
def _detect_header_row(ws):
    """Devuelve la fila del encabezado (1 o 2) detectando si hay fila de descripción."""
    row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    val = row1[MONTH_IDX] if len(row1) > MONTH_IDX else None
    return 1 if parse_hdr(val) else 2

def read_file(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Hoja1"]
    hdr_row = _detect_header_row(ws)
    header = [c.value for c in next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row))]
    items  = [list(r) for r in ws.iter_rows(min_row=hdr_row+1, values_only=True) if r[0]]
    return header, items

def read_orange(path):
    """Devuelve {item_id: {col_1based: value}} para celdas con fuente naranja."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Hoja1"]
    data_start = _detect_header_row(ws) + 1
    result = {}
    for row in ws.iter_rows(min_row=data_start):
        if not row[0].value: continue
        rid = row[0].value
        for cell in row:
            if cell.column < MONTH_IDX + 1 or not cell.value: continue
            fc = cell.font.color if cell.font else None
            if fc and fc.type == "rgb" and fc.rgb == ORANGE:
                result.setdefault(rid, {})[cell.column] = cell.value
    return result

# ══════════════════════════════════════════════════════════════════════════════
#  ALGORITMO DE CORRIMIENTO
# ══════════════════════════════════════════════════════════════════════════════
def shift_one(pcts, idx, fraction, item_max, certified_val=0.0):
    """
    Corre (pcts[idx] - certified_val) * fraction hacia adelante.
    Tope de cada mes receptor = item_max (máximo del ítem en todo el proyecto).
    Devuelve (nueva_lista, overflow).
    """
    pcts = list(pcts)
    shiftable = max(0.0, pcts[idx] - certified_val)
    amount = shiftable * fraction
    if amount < 1e-10:
        return pcts, 0.0

    pcts[idx] -= amount
    remaining  = amount
    t = idx + 1

    while remaining > 1e-10:
        if t >= len(pcts):
            return pcts, remaining          # salió del proyecto
        avail = item_max - pcts[t]
        if avail > 1e-10:
            absorb   = min(remaining, avail)
            pcts[t] += absorb
            remaining -= absorb
        t += 1

    return pcts, 0.0


def apply_delays(items, header, delays, naranja, extra_months=0):
    """
    Aplica todos los delays sobre cada ítem afectado.
    naranja: {item_id: {col_1based: certified_val}}
    Devuelve (result_items, new_header, overflow_nros).
    """
    # Extender header si corresponde
    month_names = list(header[MONTH_IDX:])
    for _ in range(extra_months):
        month_names.append(next_month_hdr(month_names[-1]))
    new_header  = list(header[:MONTH_IDX]) + month_names
    m_to_idx    = {n: i for i, n in enumerate(month_names)}

    result        = []
    overflow_nros = []

    for item in items:
        iid  = item[0]
        nro  = item[2]
        desc = item[3]
        base = [float(v or 0) for v in item[MONTH_IDX:]]
        orig = list(base)
        pcts = base + [0.0] * extra_months

        item_max = max((v for v in orig if v and v > 0), default=1.0)
        item_orange = naranja.get(iid, {})   # {col_1based: val}

        total_ovf = 0.0
        for d in delays:
            if not is_affected(nro, desc, d["motive"]):
                continue
            mn = d["month_name"]
            if mn not in m_to_idx:
                continue
            mi   = m_to_idx[mn]
            days = min(int(d["days"]), days_in(mn))
            frac = days / days_in(mn)
            # Porción certificada en ese mes (col_1based = MONTH_IDX+1 + mi)
            col1 = MONTH_IDX + 1 + mi
            cert = float(item_orange.get(col1, 0))
            pcts, ovf = shift_one(pcts, mi, frac, item_max, cert)
            total_ovf += ovf

        if total_ovf > 1e-10:
            overflow_nros.append(nro)

        new_row = list(item[:MONTH_IDX]) + [
            round(p, 6) if p > 1e-10 else None for p in pcts
        ]
        result.append(new_row)

    return result, new_header, overflow_nros


def compress_items(items, overflow_nros):
    """
    Para ítems con overflow: redistribuye el % faltante proporcionalmente
    entre los meses con valor > 0.
    """
    ovf_set = set(overflow_nros)
    result  = []
    for item in items:
        if item[2] not in ovf_set:
            result.append(item)
            continue
        pcts    = [float(v or 0) for v in item[MONTH_IDX:]]
        missing = 1.0 - sum(pcts)
        if missing > 1e-10:
            pos_sum = sum(p for p in pcts if p > 0)
            if pos_sum > 0:
                pcts = [p + missing * (p / pos_sum) if p > 0 else p for p in pcts]
        new_row = list(item[:MONTH_IDX]) + [
            round(p, 6) if p > 1e-10 else None for p in pcts
        ]
        result.append(new_row)
    return result

# ══════════════════════════════════════════════════════════════════════════════
#  GENERACIÓN DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
import math

DESC_COL_WIDTH = 48   # ancho en caracteres de la columna Descripción

def row_height(desc, font_size=8):
    """Calcula altura de fila según el largo del texto de descripción."""
    if not desc:
        return 15
    chars = len(str(desc))
    # Cada unidad de ancho ≈ 1.1 caracteres a font 8
    chars_per_line = DESC_COL_WIDTH * 1.1
    lines = max(1, math.ceil(chars / chars_per_line))
    # Cada línea ocupa ≈ font_size * 1.6 puntos; más un margen
    return max(15, int(lines * font_size * 1.7) + 4)
def _cell(ws, row, col, val=None, flt=None, fnt=None, aln=None, brd=None, fmt=None):
    c = ws.cell(row, col, val)
    if flt: c.fill      = flt
    if fnt: c.font      = fnt
    if aln: c.alignment = aln
    if brd: c.border    = brd
    if fmt: c.number_format = fmt
    return c

DIFF_THRESHOLD = 0.005   # ignorar diferencias < 0.5% (evita falsos por punto flotante)

def _diff_map(a_vals, b_vals):
    """Devuelve {idx: tag} para celdas que cambiaron más de DIFF_THRESHOLD."""
    out = {}
    for i, (a, b) in enumerate(zip(a_vals, b_vals)):
        fa = round(float(a or 0), 6)
        fb = round(float(b or 0), 6)
        if abs(fb - fa) < DIFF_THRESHOLD: continue   # diferencia insignificante → ignorar
        if   fa == 0: out[i] = "cambiado"
        elif fb == 0: out[i] = "cambiado"
        elif fb > fa: out[i] = "subio"
        else:         out[i] = "bajo"
    return out

def _write_item_row(ws, row, nro, desc, mon, prec, lbl, lbl_fill,
                    month_vals, base_fill, diff=None, sep=False):
    bdr = B_SEP if sep else B
    COL_MES1 = 6
    for ci, val in enumerate([nro, desc, mon, prec, lbl] + list(month_vals), 1):
        c = ws.cell(row, ci, val)
        c.border = bdr
        if ci == 5:                         # TIPO
            c.fill = lbl_fill; c.font = F(True,8,"FFFFFF"); c.alignment = AC
        elif ci <= 5:                       # info del ítem
            c.fill = base_fill
            c.font = F(True,8) if ci in (1,2) else F()
            c.alignment = AL if ci == 2 else AC
            if ci == 4 and val: c.number_format = "#,##0.00"
        else:                               # meses
            midx = ci - COL_MES1
            c.alignment = AC; c.font = F()
            if val is not None: c.number_format = "0%"
            tag = (diff or {}).get(midx)
            c.fill = FILLS[tag] if tag else base_fill
            if tag: c.font = F(True, 8)
    ws.row_dimensions[row].height = row_height(desc)

def _write_delta_row(ws, row, nro, vals_a, vals_b, n_meses):
    COL_MES1 = 6
    for ci in range(1, COL_MES1 + n_meses + 1):
        c = ws.cell(row, ci)
        c.border = B_SEP
        if ci == 5:
            c.value = "DIFERENCIA"; c.fill = FILLS["ldelta"]
            c.font = F(True,8,"FFFFFF"); c.alignment = AC
        elif ci < 5:
            c.fill = FILLS["dbg"]
            if ci == 1: c.value = nro; c.font = F(); c.alignment = AC
        else:
            idx = ci - COL_MES1
            va  = float(vals_a[idx] or 0) if idx < len(vals_a) else 0
            vb  = float(vals_b[idx] or 0) if idx < len(vals_b) else 0
            d   = round(vb - va, 6)
            c.alignment = AC; c.font = F(True, 8)
            if abs(d) >= DIFF_THRESHOLD:
                c.value = d; c.number_format = '+0%;-0%;""'
                c.fill  = FILLS["dpos"] if d > 0 else FILLS["dneg"]
            else:
                c.fill = FILLS["dbg"]
    ws.row_dimensions[row].height = row_height(nro)


def generate_excel(path_actual, path_contrat, epec_items, epec_header,
                   naranja, path_output, obra_name=""):
    """
    path_contrat : ruta al archivo del contratista (o None)
    epec_items   : lista de ítems con curva EPEC calculada (o None)
    epec_header  : header posiblemente extendido con meses extra (o None)
    naranja      : {item_id: {col_1based: val}}
    """
    hdr_a, items_a = read_file(path_actual)
    items_c = None
    if path_contrat and os.path.exists(path_contrat):
        _, items_c = read_file(path_contrat)

    prop_c = {r[0]: r for r in items_c}   if items_c   else {}
    prop_e = {r[0]: r for r in epec_items} if epec_items else {}

    # Header final (puede estar extendido por EPEC)
    use_hdr   = epec_header if epec_header else hdr_a
    meses     = [h for h in use_hdr[MONTH_IDX:] if h]
    N         = len(meses)
    TOTAL_COL = 5 + N        # 5 cols fijas + meses

    wb = Workbook()

    # ══ HOJA 1: COMPARATIVO ═══════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Comparativo"

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COL)
    prefix = f"{obra_name}  —  " if obra_name else ""
    t = ws.cell(1, 1, f"{prefix}ARCO  |  CURVA DE AVANCE  |  ACTUAL vs PROPUESTAS")
    t.fill = FILLS["hdr"]; t.alignment = AC
    t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws.row_dimensions[1].height = 22

    # Leyenda
    leyenda = [("ACTUAL",               FILLS["lact"],  "FFFFFF"),
               ("PROPUESTA CONTRATISTA",FILLS["lcont"], "FFFFFF"),
               ("PROPUESTA EPEC",       FILLS["lepec"], "FFFFFF"),
               ("↑ SUBIÓ",              FILLS["subio"], "000000"),
               ("↓ BAJÓ",              FILLS["bajo"],  "000000"),
               ("NUEVO / QUITADO",      FILLS["cambiado"],"000000")]
    col = 6
    for txt, flt, fc in leyenda:
        c = ws.cell(2, col, txt); c.fill = flt; c.border = B; c.alignment = AC
        c.font = Font(name="Calibri", bold=True, size=8, color=fc)
        col += 3
    ws.row_dimensions[2].height = 22

    # Encabezado columnas
    for ci, h in enumerate(["Nro","Descripción","Mon.","Precio Oficial","Tipo"] + meses, 1):
        c = ws.cell(3, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
    ws.row_dimensions[3].height = 52   # meses tienen nombres largos que envuelven

    # Anchos
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 13
    for i in range(N):
        ws.column_dimensions[get_column_letter(6+i)].width = 7

    cur = 4
    cambios_c = []
    cambios_e = []

    for item_a in items_a:
        iid  = item_a[0]
        nro  = item_a[2]; desc = item_a[3]
        mon  = item_a[5]; prec = item_a[9]
        ma   = [item_a[MONTH_IDX+i] if MONTH_IDX+i < len(item_a) else None for i in range(N)]

        # Propuesta contratista
        rc   = prop_c.get(iid)
        mc   = [rc[MONTH_IDX+i] if rc and MONTH_IDX+i < len(rc) else None for i in range(N)]

        # Propuesta EPEC
        re   = prop_e.get(iid)
        me   = [re[MONTH_IDX+i] if re and MONTH_IDX+i < len(re) else None for i in range(N)]

        # Detectar cambios para resumen
        if rc and any(round(float(a or 0),6) != round(float(c or 0),6) for a,c in zip(ma,mc)):
            cambios_c.append((nro, desc))
        if re and any(round(float(a or 0),6) != round(float(e or 0),6) for a,e in zip(ma,me)):
            cambios_e.append((nro, desc))

        # Fila ACTUAL
        _write_item_row(ws, cur, nro, desc, mon, prec,
                        "ACTUAL", FILLS["lact"], ma, FILLS["actual"])
        cur += 1

        # Fila CONTRATISTA + delta
        if rc:
            dm = _diff_map(ma, mc)
            _write_item_row(ws, cur, nro, desc, mon, prec,
                            "CONTRATISTA", FILLS["lcont"], mc, FILLS["contrat"], dm)
            cur += 1
            _write_delta_row(ws, cur, nro, ma, mc, N)
            cur += 1

        # Fila EPEC + delta
        if re:
            dm = _diff_map(ma, me)
            _write_item_row(ws, cur, nro, desc, mon, prec,
                            "PROPUESTA EPEC", FILLS["lepec"], me, FILLS["epec"], dm,
                            sep=not rc)   # sep solo si es la última fila del grupo
            cur += 1
            _write_delta_row(ws, cur, nro, ma, me, N)
            cur += 1

        # Si solo hay fila ACTUAL (sin propuestas), poner separador
        if not rc and not re:
            for ci2 in range(1, TOTAL_COL+1):
                ws.cell(cur-1, ci2).border = B_SEP

    ws.freeze_panes = ws.cell(4, 6)

    # ══ HOJA: EPEC vs CONTRATISTA (solo si ambas propuestas están presentes) ══
    if prop_c and prop_e:
        ws_ce = wb.create_sheet("EPEC vs Contratista")

        # Anchos
        ws_ce.column_dimensions["A"].width = 7
        ws_ce.column_dimensions["B"].width = 48
        ws_ce.column_dimensions["C"].width = 5
        ws_ce.column_dimensions["D"].width = 14
        ws_ce.column_dimensions["E"].width = 13
        for i in range(N):
            ws_ce.column_dimensions[get_column_letter(6+i)].width = 7

        # Título
        ws_ce.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COL)
        t_ce = ws_ce.cell(1, 1, f"{prefix}ARCO  |  PROPUESTA EPEC vs PROPUESTA CONTRATISTA")
        t_ce.fill = FILLS["hdr"]; t_ce.alignment = AC
        t_ce.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        ws_ce.row_dimensions[1].height = 22

        # Leyenda
        leyenda_ce = [
            ("PROPUESTA EPEC",        FILLS["lepec"],    "FFFFFF"),
            ("PROPUESTA CONTRATISTA", FILLS["lcont"],    "FFFFFF"),
            ("CONTRAT. PIDE MÁS",     FILLS["bajo"],     "000000"),
            ("CONTRAT. PIDE MENOS",   FILLS["subio"],    "000000"),
            ("NUEVO / QUITADO",       FILLS["cambiado"], "000000"),
        ]
        col_l = 6
        for txt, flt, fc in leyenda_ce:
            c = ws_ce.cell(2, col_l, txt); c.fill = flt; c.border = B; c.alignment = AC
            c.font = Font(name="Calibri", bold=True, size=8, color=fc)
            col_l += 3
        ws_ce.row_dimensions[2].height = 22

        # Encabezado
        for ci, h in enumerate(["Nro","Descripción","Mon.","Precio Oficial","Tipo"] + meses, 1):
            c = ws_ce.cell(3, ci, h); c.fill = FILLS["hdr"]
            c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
        ws_ce.row_dimensions[3].height = 52

        cur_ce = 4
        for item_a in items_a:
            iid  = item_a[0]
            nro  = item_a[2]; desc = item_a[3]
            mon  = item_a[5]; prec = item_a[9]

            rc_ce = prop_c.get(iid)
            re_ce = prop_e.get(iid)
            if not rc_ce and not re_ce:
                continue

            mc_ce = [rc_ce[MONTH_IDX+i] if rc_ce and MONTH_IDX+i < len(rc_ce) else None for i in range(N)]
            me_ce = [re_ce[MONTH_IDX+i] if re_ce and MONTH_IDX+i < len(re_ce) else None for i in range(N)]

            # Diff map aplicado sobre la fila CONTRATISTA:
            # Rojo   = contratista pide MÁS que EPEC
            # Verde  = contratista pide MENOS que EPEC
            diff_ce = {}
            for i, (vc2, ve2) in enumerate(zip(mc_ce, me_ce)):
                fc2 = round(float(vc2 or 0), 6)
                fe2 = round(float(ve2 or 0), 6)
                if abs(fc2 - fe2) < DIFF_THRESHOLD: continue
                if   fc2 == 0: diff_ce[i] = "cambiado"
                elif fe2 == 0: diff_ce[i] = "cambiado"
                elif fc2 > fe2: diff_ce[i] = "bajo"   # contratista pide más → rojo
                else:           diff_ce[i] = "subio"  # contratista pide menos → verde

            # Fila EPEC (base, sin highlight)
            _write_item_row(ws_ce, cur_ce, nro, desc, mon, prec,
                            "PROPUESTA EPEC", FILLS["lepec"], me_ce, FILLS["epec"])
            cur_ce += 1

            # Fila CONTRATISTA (con highlighting respecto a EPEC)
            _write_item_row(ws_ce, cur_ce, nro, desc, mon, prec,
                            "CONTRATISTA", FILLS["lcont"], mc_ce, FILLS["contrat"], diff_ce)
            cur_ce += 1

            # Fila DELTA: CONTRATISTA - EPEC (positivo = contratista pide más)
            _write_delta_row(ws_ce, cur_ce, nro, me_ce, mc_ce, N)
            cur_ce += 1

        ws_ce.freeze_panes = ws_ce.cell(4, 6)

    # ══ HOJA 2: RESUMEN DE CAMBIOS ════════════════════════════════════════════
    ws2 = wb.create_sheet("Resumen de Cambios")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 22

    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value     = f"{prefix}RESUMEN DE CAMBIOS"
    t2.fill      = FILLS["hdr"]; t2.alignment = AC
    t2.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws2.row_dimensions[1].height = 20

    for ci, h in enumerate(["Nro","Descripción","Propuesta"], 1):
        c = ws2.cell(2, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B

    rr = 3
    for nro, desc in cambios_c:
        a=ws2.cell(rr,1,nro);   a.border=B; a.alignment=AC; a.font=F()
        b=ws2.cell(rr,2,desc);  b.border=B; b.alignment=AL; b.font=F()
        c=ws2.cell(rr,3,"Contratista"); c.border=B; c.alignment=AC
        c.font=F(True,8,"FFFFFF"); c.fill=FILLS["lcont"]
        flt = fill("FFFFFF" if rr%2==0 else "F2F2F2")
        a.fill=flt; b.fill=flt
        ws2.row_dimensions[rr].height = row_height(desc); rr+=1

    for nro, desc in cambios_e:
        a=ws2.cell(rr,1,nro);   a.border=B; a.alignment=AC; a.font=F()
        b=ws2.cell(rr,2,desc);  b.border=B; b.alignment=AL; b.font=F()
        c=ws2.cell(rr,3,"EPEC"); c.border=B; c.alignment=AC
        c.font=F(True,8,"FFFFFF"); c.fill=FILLS["lepec"]
        flt = fill("FFFFFF" if rr%2==0 else "F2F2F2")
        a.fill=flt; b.fill=flt
        ws2.row_dimensions[rr].height = row_height(desc); rr+=1

    # ══ HOJA 3: VERIFICACIÓN CERTIFICADOS ════════════════════════════════════
    ws3 = wb.create_sheet("Verificación Certificados")
    for col_l, w in zip("ABCDEFG", [8,52,22,18,18,18,16]):
        ws3.column_dimensions[col_l].width = w

    violations = []; ok_list = []
    for item_a in items_a:
        iid = item_a[0]; nro = item_a[2]; desc = item_a[3]
        orange_cols = naranja.get(iid, {})
        if not orange_cols: continue
        for col1, val_a in orange_cols.items():
            mes_name = hdr_a[col1-1] if col1-1 < len(hdr_a) else ""
            rc2 = prop_c.get(iid)
            re2 = prop_e.get(iid)
            vc  = rc2[col1-1] if rc2 and col1-1 < len(rc2) else None
            ve  = re2[col1-1] if re2 and col1-1 < len(re2) else None
            fa  = round(float(val_a or 0), 8)
            fc2 = round(float(vc or 0),    8) if vc is not None else fa
            fe2 = round(float(ve or 0),    8) if ve is not None else fa
            entry = dict(nro=nro, desc=desc, mes=mes_name,
                         va=val_a, vc=vc, ve=ve)
            if fa != fc2 or fa != fe2:
                violations.append(entry)
            else:
                ok_list.append(entry)

    ws3.merge_cells("A1:G1")
    b1 = ws3["A1"]
    if violations:
        b1.value = f"ALERTA: {len(violations)} celda(s) certificada(s) modificadas"
        b1.fill  = FILLS["warn_b"]
    else:
        b1.value = f"VERIFICACIÓN OK — {len(ok_list)} celdas certificadas sin modificar"
        b1.fill  = FILLS["ok_b"]
    b1.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    b1.alignment = AC; ws3.row_dimensions[1].height = 24

    for ci, h in enumerate(["Nro","Descripción","Mes","Actual","Contratista","EPEC","Estado"], 1):
        c = ws3.cell(2, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True,9,"FFFFFF"); c.alignment = AC; c.border = B
    ws3.row_dimensions[2].height = 18

    rv = 3
    for e in violations + ok_list:
        is_v = e in violations
        flt  = FILLS["viol_row"] if is_v else FILLS["ok_row"]
        est  = "MODIFICADO" if is_v else "OK"
        fc_e = Font(name="Calibri", bold=True, size=8,
                    color=("C00000" if is_v else "375623"))
        for ci, v in enumerate([e["nro"],e["desc"],e["mes"],
                                 e["va"],e["vc"],e["ve"],est], 1):
            c = ws3.cell(rv, ci, v); c.fill = flt; c.border = B
            c.alignment = AL if ci == 2 else AC
            if ci == 7: c.font = fc_e
            elif ci in (4,5,6):
                c.font = F(True,8)
                if v is not None: c.number_format = "0%"
            else: c.font = F()
        ws3.row_dimensions[rv].height = row_height(e["desc"]); rv += 1

    ws3.freeze_panes = ws3["A3"]

    # ══ HOJA: CURVAS ══════════════════════════════════════════════════════════
    def calc_cumul(item_list, n):
        """Suma precio_oficial × % por mes, acumula. Devuelve lista de n valores."""
        monthly = [0.0] * n
        for item in item_list:
            precio = float(item[9] or 0)
            for i in range(n):
                idx = MONTH_IDX + i
                pct = float(item[idx] or 0) if idx < len(item) else 0.0
                monthly[i] += precio * pct
        cumul = []
        acc = 0.0
        for v in monthly:
            acc += v
            cumul.append(round(acc, 2))
        return cumul

    wc = wb.create_sheet("Curvas")

    # ── Calcular acumulados ────────────────────────────────────────────────────
    cumul_a = calc_cumul(items_a, N)
    total_a = cumul_a[-1] if cumul_a else 1.0

    cumul_c = calc_cumul(list(prop_c.values()), N) if prop_c else None
    cumul_e = calc_cumul(list(prop_e.values()), N) if prop_e else None

    # ── Tabla de datos ─────────────────────────────────────────────────────────
    # Fila 1: título
    wc.merge_cells("A1:H1")
    tc = wc["A1"]
    tc.value = f"{prefix}CURVAS DE AVANCE ACUMULADO"
    tc.fill = FILLS["hdr"]; tc.alignment = AC
    tc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    wc.row_dimensions[1].height = 22

    # Fila 2: encabezados
    headers_c = ["Mes",
                 "Actual ($)", "Actual (%)",
                 "Propuesta EPEC ($)", "Propuesta EPEC (%)",
                 "Contratista ($)", "Contratista (%)"]
    for ci, h in enumerate(headers_c, 1):
        c = wc.cell(2, ci, h); c.fill = FILLS["hdr"]
        c.font = F(True, 9, "FFFFFF"); c.alignment = AC; c.border = B
    wc.row_dimensions[2].height = 28

    # Anchos
    wc.column_dimensions["A"].width = 22
    for col_l in "BCDEFG":
        wc.column_dimensions[col_l].width = 18

    # Filas de datos (fila 3 en adelante)
    for i, mes in enumerate(meses):
        r = i + 3
        flt = fill("FFFFFF" if i % 2 == 0 else "F2F2F2")

        # Mes
        c = wc.cell(r, 1, mes); c.fill = flt; c.border = B
        c.font = F(); c.alignment = AC

        # Actual
        va = cumul_a[i]
        c = wc.cell(r, 2, va); c.fill = flt; c.border = B
        c.font = F(); c.alignment = AC; c.number_format = '#,##0'
        pct_a = va / total_a if total_a else 0
        c = wc.cell(r, 3, pct_a); c.fill = flt; c.border = B
        c.font = F(); c.alignment = AC; c.number_format = '0%'

        # EPEC
        if cumul_e:
            ve = cumul_e[i]
            total_e = cumul_e[-1] if cumul_e[-1] else 1.0
            c = wc.cell(r, 4, ve); c.fill = flt; c.border = B
            c.font = F(); c.alignment = AC; c.number_format = '#,##0'
            c = wc.cell(r, 5, ve / total_e); c.fill = flt; c.border = B
            c.font = F(); c.alignment = AC; c.number_format = '0%'

        # Contratista
        if cumul_c:
            vc = cumul_c[i]
            total_c = cumul_c[-1] if cumul_c[-1] else 1.0
            c = wc.cell(r, 6, vc); c.fill = flt; c.border = B
            c.font = F(); c.alignment = AC; c.number_format = '#,##0'
            c = wc.cell(r, 7, vc / total_c); c.fill = flt; c.border = B
            c.font = F(); c.alignment = AC; c.number_format = '0%'

        wc.row_dimensions[r].height = 15

    n_data = N  # filas de datos

    # ── Gráfico de líneas ─────────────────────────────────────────────────────
    chart = LineChart()
    chart.title        = "Curvas de Avance Acumulado"
    chart.style        = 10
    chart.height       = 18
    chart.width        = 32

    # Eje Y — mostrar en millones con separador de miles
    # Formato: si el monto máximo supera 1.000M, mostrar en millones
    max_val = max(cumul_a) if cumul_a else 1
    if max_val >= 1_000_000_000:
        # Mostrar en millones: 1.500.000.000 → "$ 1.500 M"
        chart.y_axis.numFmt  = '#,##0,,"M"'
        chart.y_axis.title   = "Monto acumulado (en millones $)"
    elif max_val >= 1_000_000:
        chart.y_axis.numFmt  = '#,##0,"k"'
        chart.y_axis.title   = "Monto acumulado (en miles $)"
    else:
        chart.y_axis.numFmt  = '#,##0'
        chart.y_axis.title   = "Monto acumulado ($)"

    chart.y_axis.delete        = False   # asegurar que el eje se muestre
    chart.y_axis.crosses       = "autoZero"
    chart.y_axis.crossAx       = 100    # ID del eje X
    chart.x_axis.title         = "Mes"
    chart.x_axis.tickLblSkip   = 1      # mostrar todas las etiquetas
    chart.x_axis.tickMarkSkip  = 1
    chart.x_axis.noMultiLvlLbl = True   # una sola línea de etiquetas

    # Etiquetas del eje X (meses)
    cats = Reference(wc, min_col=1, min_row=3, max_row=2 + n_data)
    chart.set_categories(cats)

    # Serie Actual
    ser_a = Reference(wc, min_col=2, min_row=2, max_row=2 + n_data)
    chart.add_data(ser_a, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill  = "2E75B6"
    chart.series[0].graphicalProperties.line.width      = 20000
    chart.series[0].smooth = True

    # Serie EPEC
    if cumul_e:
        ser_e = Reference(wc, min_col=4, min_row=2, max_row=2 + n_data)
        chart.add_data(ser_e, titles_from_data=True)
        idx_e = len(chart.series) - 1
        chart.series[idx_e].graphicalProperties.line.solidFill = "7030A0"
        chart.series[idx_e].graphicalProperties.line.width     = 20000
        chart.series[idx_e].smooth = True

    # Serie Contratista
    if cumul_c:
        ser_c = Reference(wc, min_col=6, min_row=2, max_row=2 + n_data)
        chart.add_data(ser_c, titles_from_data=True)
        idx_c = len(chart.series) - 1
        chart.series[idx_c].graphicalProperties.line.solidFill = "70AD47"
        chart.series[idx_c].graphicalProperties.line.width     = 20000
        chart.series[idx_c].smooth = True

    # Ubicar el gráfico debajo de la tabla
    chart_row = n_data + 5
    wc.add_chart(chart, f"A{chart_row}")

    wb.save(path_output)


# ══════════════════════════════════════════════════════════════════════════════
#  INTERFAZ GRÁFICA
# ══════════════════════════════════════════════════════════════════════════════
def _short_month(full_name):
    """'31 de marzo de 2026' → 'Marzo 2026  (31d)'"""
    p = parse_hdr(full_name)
    if not p: return full_name
    m, y = p
    nombre = MESES_REV[m].capitalize()
    d = calendar.monthrange(y, m)[1]
    return f"{nombre} {y}  ({d}d)"


class DelayRow(tk.Frame):
    def __init__(self, parent, months, on_delete):
        """
        months: lista de nombres completos de meses (ej. '31 de marzo de 2026')
        El dropdown muestra nombres cortos; internamente se guarda el nombre completo.
        """
        super().__init__(parent, bg="#F5F0FF", pady=2)
        self.motive = tk.StringVar(value=MOTIVOS[0])
        self.days   = tk.StringVar(value="15")
        self._full_months   = list(months)                         # nombres completos
        self._display_months = [_short_month(m) for m in months]  # nombres cortos
        self._d2f = dict(zip(self._display_months, self._full_months))

        # Selector de mes usa nombres CORTOS para claridad
        self._month_display = tk.StringVar(
            value=self._display_months[0] if self._display_months else "")

        ttk.Combobox(self, textvariable=self.motive, values=MOTIVOS,
                     width=32, state="readonly").pack(side="left", padx=(4,4))

        self.cb_mes = ttk.Combobox(self, textvariable=self._month_display,
                                   values=self._display_months, width=22, state="readonly")
        self.cb_mes.pack(side="left", padx=(0,4))

        vcmd = (self.register(lambda s: s == "" or (s.isdigit() and 1 <= int(s) <= 31)), "%P")
        tk.Entry(self, textvariable=self.days, width=4,
                 validate="key", validatecommand=vcmd,
                 font=("Calibri",9), relief="solid", bd=1).pack(side="left")

        tk.Label(self, text=" días", bg="#F5F0FF",
                 font=("Calibri",8,"italic"), fg="#666").pack(side="left")

        tk.Button(self, text="✕", command=on_delete,
                  bg="#DDD", relief="flat", width=2,
                  cursor="hand2", font=("Calibri",9,"bold")).pack(side="left", padx=(8,4))

    def update_months(self, months):
        self._full_months    = list(months)
        self._display_months = [_short_month(m) for m in months]
        self._d2f            = dict(zip(self._display_months, self._full_months))
        self.cb_mes["values"] = self._display_months
        if self._display_months and self._month_display.get() not in self._display_months:
            self._month_display.set(self._display_months[0])

    def get(self):
        disp = self._month_display.get()
        full = self._d2f.get(disp, disp)   # si no mapea, usa el valor directo
        return {"motive":     self.motive.get(),
                "month_name": full,
                "days":       int(self.days.get() or 1)}


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ARCO — Análisis y Revisión de Curvas de Obra  v2.0")
        self.resizable(False, False)
        self.configure(bg="#1F3864")

        self.v_obra    = tk.StringVar()
        self.v_actual  = tk.StringVar()
        self.v_contrat = tk.StringVar()
        self.v_salida  = tk.StringVar()
        self.v_epec    = tk.BooleanVar(value=False)
        self.v_status  = tk.StringVar(value="Seleccioná el archivo de Curva Actual para comenzar.")

        self._months    = []
        self._delay_rows = []

        self._build()
        self._center()
        # Actualizar nombre de archivo sugerido cuando cambia el nombre de obra
        self.v_obra.trace_add("write", lambda *_: self._refresh_salida())

    def _center(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - self.winfo_width())  // 2
        y = (self.winfo_screenheight() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    # ── CONSTRUCCIÓN DE UI ────────────────────────────────────────────────────
    def _build(self):
        tk.Label(self, text="ARCO",
                 bg="#1F3864", fg="white",
                 font=("Calibri",15,"bold")).pack(pady=(16,2))
        tk.Label(self, text="Curva de Avance",
                 bg="#1F3864", fg="#AAC4E8", font=("Calibri",9)).pack(pady=(0,10))

        self.card = tk.Frame(self, bg="white")
        self.card.pack(padx=20, fill="x")

        # Nombre de obra
        self._sec("IDENTIFICACIÓN DE OBRA")
        obra_row = tk.Frame(self.card, bg="white")
        obra_row.pack(fill="x", padx=16, pady=8)
        tk.Label(obra_row, text="Nombre de la obra", bg="white", fg="#444",
                 font=("Calibri",9,"bold"), anchor="w").pack(fill="x")
        tk.Entry(obra_row, textvariable=self.v_obra, width=55,
                 font=("Calibri",10), relief="solid", bd=1,
                 fg="#1F3864").pack(fill="x", pady=(4,0), ipady=5)
        tk.Label(obra_row, text="Aparece en el título del Excel y en el nombre del archivo",
                 bg="white", fg="#999", font=("Calibri",8,"italic")).pack(anchor="w")

        tk.Frame(self.card, bg="#444", height=2).pack(fill="x", pady=4)

        # Archivos
        self._sec("ARCHIVOS")
        self._file_row("Curva Actual  (requerida)",     self.v_actual,  self._browse_actual,  "#2E75B6")
        tk.Frame(self.card, bg="#E0E0E0", height=1).pack(fill="x", padx=16)
        self._file_row("Propuesta Contratista  (opcional)", self.v_contrat, self._browse_contrat, "#70AD47")

        tk.Frame(self.card, bg="#444", height=2).pack(fill="x", pady=6)

        # Sección EPEC
        epec_toggle = tk.Frame(self.card, bg="white")
        epec_toggle.pack(fill="x", padx=16, pady=(2,4))
        tk.Checkbutton(epec_toggle,
                       text="  GENERAR PROPUESTA EPEC  (prórroga por motivos)",
                       variable=self.v_epec, bg="white", fg="#7030A0",
                       font=("Calibri",10,"bold"), cursor="hand2",
                       activebackground="white", selectcolor="white",
                       command=self._toggle_epec).pack(side="left")

        # Contenido EPEC (oculto hasta activar)
        self.epec_body = tk.Frame(self.card, bg="#F5F0FF")

        # Encabezado tabla
        hf = tk.Frame(self.epec_body, bg="#EDE7F6")
        hf.pack(fill="x", padx=8, pady=(4,2))
        for txt, w in [("Motivo de prórroga",33),("Mes afectado",22),("Días",6)]:
            tk.Label(hf, text=txt, bg="#EDE7F6", fg="#4A148C",
                     font=("Calibri",8,"bold"), width=w, anchor="w").pack(side="left", padx=4)

        self.delay_container = tk.Frame(self.epec_body, bg="#F5F0FF")
        self.delay_container.pack(fill="x", padx=4)

        tk.Button(self.epec_body, text="＋  Agregar motivo",
                  command=self._add_delay,
                  bg="#7030A0", fg="white", relief="flat",
                  font=("Calibri",9,"bold"), cursor="hand2",
                  padx=10, pady=4).pack(anchor="w", padx=8, pady=(6,8))

        tk.Frame(self.card, bg="#444", height=2).pack(fill="x", pady=4)

        # Salida
        self._file_row("Guardar comparativo como...", self.v_salida,
                       self._browse_salida, "#595959", save=True)

        # Botón generar
        bf = tk.Frame(self, bg="#1F3864")
        bf.pack(pady=12)
        self.btn = tk.Button(bf, text="  ▶  GENERAR  ",
                             command=self._run,
                             bg="#2E75B6", fg="white",
                             font=("Calibri",13,"bold"),
                             relief="flat", cursor="hand2",
                             padx=30, pady=10,
                             activebackground="#1F3864", activeforeground="white")
        self.btn.pack()

        self.pbar = ttk.Progressbar(self, mode="indeterminate", length=420)
        self.pbar.pack(pady=(0,4))

        tk.Label(self, textvariable=self.v_status,
                 bg="#1F3864", fg="#AAC4E8",
                 font=("Calibri",8), wraplength=440).pack(pady=(0,14))

    def _sec(self, title):
        f = tk.Frame(self.card, bg="#E8EAF6")
        f.pack(fill="x")
        tk.Label(f, text=f"  {title}", bg="#E8EAF6", fg="#1F3864",
                 font=("Calibri",9,"bold"), anchor="w").pack(fill="x", pady=4)

    def _file_row(self, label, var, cmd, color, save=False):
        row = tk.Frame(self.card, bg="white")
        row.pack(fill="x", padx=16, pady=8)
        tk.Label(row, text=label, bg="white", fg="#444",
                 font=("Calibri",9,"bold"), anchor="w").pack(fill="x")
        inner = tk.Frame(row, bg="white")
        inner.pack(fill="x", pady=(3,0))
        tk.Entry(inner, textvariable=var, width=52,
                 font=("Calibri",8), relief="solid", bd=1,
                 fg="#555", state="readonly").pack(side="left", fill="x",
                                                   expand=True, ipady=4)
        tk.Button(inner, text="Examinar...", command=cmd,
                  bg=color, fg="white", relief="flat",
                  font=("Calibri",8,"bold"), cursor="hand2",
                  padx=8, pady=4,
                  activebackground="#1F3864", activeforeground="white"
                  ).pack(side="left", padx=(6,0))

    # ── EVENTOS UI ────────────────────────────────────────────────────────────
    def _toggle_epec(self):
        if self.v_epec.get():
            self.epec_body.pack(fill="x", padx=16, pady=(0,6),
                                in_=self.card,
                                before=self.card.winfo_children()[-2])
            if not self._delay_rows:
                self._add_delay()
        else:
            self.epec_body.pack_forget()

    def _add_delay(self):
        if not self._months:
            messagebox.showwarning("Sin archivo",
                                   "Cargá primero el archivo de Curva Actual.")
            self.v_epec.set(False)
            self.epec_body.pack_forget()
            return

        def del_row(r):
            r.destroy()
            if r in self._delay_rows:
                self._delay_rows.remove(r)

        row = DelayRow(self.delay_container, self._months,
                       on_delete=lambda: None)
        # Wire delete correctly after creation
        for w in row.winfo_children():
            if isinstance(w, tk.Button) and w["text"] == "✕":
                w.config(command=lambda r=row: del_row(r))
                break
        row.pack(fill="x", pady=1)
        self._delay_rows.append(row)

    def _browse_actual(self):
        p = filedialog.askopenfilename(
            title="Curva Actual",
            filetypes=[("Excel","*.xlsx *.xlsm"),("Todos","*.*")])
        if not p: return
        self.v_actual.set(p)
        self._load_months(p)
        self._auto_salida()

    def _browse_contrat(self):
        p = filedialog.askopenfilename(
            title="Propuesta Contratista",
            filetypes=[("Excel","*.xlsx *.xlsm"),("Todos","*.*")])
        if p: self.v_contrat.set(p)

    def _browse_salida(self):
        p = filedialog.asksaveasfilename(
            title="Guardar como",
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile="ARCO_Curva_Avance.xlsx")
        if p: self.v_salida.set(p)

    def _make_fname(self):
        obra = self.v_obra.get().strip()
        safe = "".join(c if c.isalnum() or c in "_-" else "_" for c in obra).strip("_")
        return f"ARCO_{safe}.xlsx" if safe else "ARCO_Curva_Avance.xlsx"

    def _auto_salida(self):
        if self.v_salida.get(): return
        base = self.v_actual.get()
        if base:
            self.v_salida.set(os.path.join(str(Path(base).parent), self._make_fname()))

    def _refresh_salida(self):
        """Actualiza el nombre de archivo cuando cambia el nombre de obra."""
        current = self.v_salida.get()
        base = self.v_actual.get()
        if not base: return
        folder = str(Path(base).parent)
        # Solo actualizar si el archivo sugerido sigue en la misma carpeta
        if current and str(Path(current).parent) == folder:
            self.v_salida.set(os.path.join(folder, self._make_fname()))

    def _load_months(self, path):
        try:
            hdr, _ = read_file(path)
            self._months = [h for h in hdr[MONTH_IDX:] if h]
            for r in self._delay_rows:
                r.update_months(self._months)
            self._set_st(f"Archivo cargado — {len(self._months)} meses detectados.")
        except Exception as ex:
            messagebox.showerror("Error al leer", str(ex))

    def _set_st(self, msg):
        self.after(0, lambda: self.v_status.set(msg))

    # ── DIÁLOGO OVERFLOW (thread-safe) ────────────────────────────────────────
    def _ask_overflow(self):
        result = [None]
        evt    = threading.Event()
        def ask():
            result[0] = messagebox.askyesno(
                "Corrimiento supera fin del proyecto",
                "El corrimiento de algunos ítems alcanza\n"
                "más allá del último mes planificado.\n\n"
                "¿Querés EXTENDER el plazo del proyecto?\n\n"
                "  SÍ  →  Se agregan meses al final\n"
                "  NO  →  Se comprimen las actividades finales\n"
                "          (redistribución proporcional)")
            evt.set()
        self.after(0, ask)
        evt.wait()
        return result[0]

    # ── GENERACIÓN ────────────────────────────────────────────────────────────
    def _run(self):
        pa  = self.v_actual.get().strip()
        pc  = self.v_contrat.get().strip()
        ps  = self.v_salida.get().strip()
        use = self.v_epec.get()

        if not pa:
            messagebox.showwarning("Falta archivo", "Cargá la Curva Actual."); return
        if not pc and not use:
            messagebox.showwarning("Falta configuración",
                                   "Cargá la Propuesta del Contratista\n"
                                   "o activá la Propuesta EPEC."); return
        if not ps:
            messagebox.showwarning("Falta destino",
                                   "Indicá dónde guardar el archivo."); return

        delays = []
        if use:
            delays = [r.get() for r in self._delay_rows if r.winfo_exists()]
            if not delays:
                messagebox.showwarning("Sin motivos",
                                       "Agregá al menos un motivo de prórroga."); return

        self.btn.config(state="disabled")
        self.pbar.start(12)

        def worker():
            try:
                self._set_st("Leyendo archivos...")
                hdr_a, items_a = read_file(pa)
                naranja        = read_orange(pa)

                ep_items  = None
                ep_header = None

                if use:
                    self._set_st("Calculando corrimientos EPEC...")
                    ep_items, ep_hdr, ovf = apply_delays(
                        items_a, hdr_a, delays, naranja, extra_months=0)

                    if ovf:
                        extend = self._ask_overflow()
                        if extend:
                            self._set_st("Extendiendo plazo...")
                            # Intentar con hasta 6 meses extra
                            for extra in [2, 4, 6]:
                                ep_items, ep_hdr, ovf = apply_delays(
                                    items_a, hdr_a, delays, naranja, extra_months=extra)
                                if not ovf:
                                    break
                        else:
                            self._set_st("Comprimiendo actividades...")
                            ep_items = compress_items(ep_items, ovf)

                    ep_header = ep_hdr

                self._set_st("Generando Excel...")
                generate_excel(
                    path_actual  = pa,
                    path_contrat = pc if pc
else None,
                    epec_items   = ep_items,
                    epec_header  = ep_header,
                    naranja      = naranja,
                    path_output  = ps,
                    obra_name    = self.v_obra.get().strip(),
                )

                self.after(0, self.pbar.stop)
                self.after(0, lambda: self.btn.config(state="normal"))
                self._set_st("¡Listo!")
                self.after(0, lambda: messagebox.showinfo(
                    "ARCO generado",
                    f"Archivo guardado correctamente en:\n{ps}"))
                try: os.startfile(ps)
                except Exception: pass

            except Exception as ex:
                msg = str(ex)
                tb  = __import__('traceback').format_exc()
                if "Permission denied" in msg or "errno 13" in msg.lower():
                    msg = ("No se puede guardar el archivo porque está abierto en Excel.\n\n"
                           "Cerrá el archivo y volvé a intentarlo.")
                    tb = ""
                self.after(0, self.pbar.stop)
                self.after(0, lambda: self.btn.config(state="normal"))
                self._set_st(f"Error: {msg}")
                self.after(0, lambda m=msg, t=tb: messagebox.showerror(
                    "Error", m + ("\n\n" + t if t else "")))

        threading.Thread(target=worker, daemon=True).start()


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    App().mainloop()
